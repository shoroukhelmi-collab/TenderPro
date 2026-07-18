"""Excel ingestion utilities for TenderPro."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import BinaryIO, Iterable

import pandas as pd
from openpyxl import load_workbook
from modules.mapping import REQUIRED_COLUMNS, detect_columns

COMMERCIAL_COLUMNS = (*REQUIRED_COLUMNS, "package", "supplier", "file_name", "worksheet")
HEADER_SCAN_ROWS = 60
MIN_HEADER_ROWS = 1
MAX_HEADER_ROWS = 5
TOTAL_PATTERNS = re.compile(r"\b(sub\s*total|subtotal|grand\s*total|total|summary|carry\s*forward)\b", re.I)
NOTE_PATTERNS = re.compile(r"\b(note|notes|description of works|general requirement|preamble|specification|specifications|tender|boq|bill of quantities|method of measurements?|methods of measurements?|scope of work)\b", re.I)
FILENAME_NOISE_PATTERNS = re.compile(
    r"\b(rev(?:ision)?\.?\s*[a-z0-9]+|r\d+|v\d+|final|draft|copy|priced|quote|quotation|boq|bill of quantities|tender|commercial)\b",
    re.I,
)
NON_PACKAGE_HEADINGS = re.compile(r"^\s*(method of measurements?|methods of measurements?|general notes?|specifications?|scope of work)\s*$", re.I)
PACKAGE_FROM_HEADINGS = "Section headings"
SECTION_PATTERNS = re.compile(r"\b(section|package|trade|bill\s*no|part|division|schedule)\b", re.I)
SAMPLE_SUPPLIERS: dict[str, list[dict[str, object]]] = {
    "Supplier_1.xlsx": [
        {"Item No": "1.01", "Item Description": "Concrete footing", "UOM": "m3", "Qty": 25, "Rate": 95},
        {"Item No": "1.02", "Item Description": "Steel reinforcement", "UOM": "kg", "Qty": 1200, "Rate": 1.8},
        {"Item No": "1.03", "Item Description": "Cable tray", "UOM": "m", "Qty": 80, "Rate": None},
    ],
    "Supplier_2.xlsx": [
        {"Code": "1.01", "Description": "Concrete footing", "Unit": "m3", "Quantity": 25, "Unit Price": 90},
        {"Code": "1.02", "Description": "Steel reinforcement", "Unit": "kg", "Quantity": 1200, "Unit Price": 1.95},
        {"Code": "1.03", "Description": "Cable tray", "Unit": "m", "Quantity": 80, "Unit Price": 22},
    ],
}


@dataclass
class WorksheetReview:
    """Detected worksheet metadata shown before comparison generation."""

    sheet_name: str
    header_row: int
    column_mapping: dict[str, str | None]
    imported_rows: int
    excluded_rows: int
    columns: list[str]
    package_source: str = PACKAGE_FROM_HEADINGS
    section_headings: list[str] = field(default_factory=list)


@dataclass
class WorkbookReview:
    """Detected workbook metadata shown before comparison generation."""

    file_name: str
    supplier_name: str
    worksheet_names: list[str]
    selected_worksheet: str
    header_row: int
    column_mapping: dict[str, str | None]
    imported_rows: int
    columns: list[str]
    excluded_rows: int = 0
    package_source: str = PACKAGE_FROM_HEADINGS
    section_headings: list[str] = field(default_factory=list)
    worksheets: list[WorksheetReview] = field(default_factory=list)


def sample_workbooks(directory: str | Path = "uploads") -> list[Path]:
    """Create small generic sample workbooks for local demos."""
    output_dir = Path(directory)
    output_dir.mkdir(parents=True, exist_ok=True)
    paths: list[Path] = []
    for filename, rows in SAMPLE_SUPPLIERS.items():
        path = output_dir / filename
        frame = pd.DataFrame(rows)
        if "Rate" in frame.columns:
            frame["Amount"] = frame["Qty"] * frame["Rate"]
        if "Unit Price" in frame.columns:
            frame["Total"] = frame["Quantity"] * frame["Unit Price"]
        frame.to_excel(path, index=False)
        paths.append(path)
    return paths


def inspect_excel_file(file: str | Path | BinaryIO, supplier_name: str | None = None) -> WorkbookReview:
    """Inspect all sheets in a workbook independently."""
    name = getattr(file, "name", str(file))
    detected_supplier = supplier_name or detect_supplier_name(name)
    _rewind(file)
    excel = pd.ExcelFile(file, engine="openpyxl")
    worksheet_reviews: list[WorksheetReview] = []
    for sheet in excel.sheet_names:
        header_row, columns, mapping = _detect_header(excel, sheet)
        package_col = _find_package_column(columns)
        if package_col:
            mapping["package"] = package_col
        rows, excluded_count, headings = _read_sheet(file, detected_supplier, sheet, header_row, mapping)
        worksheet_reviews.append(WorksheetReview(sheet, header_row + 1, mapping, len(rows), excluded_count, columns, _package_source(mapping), headings[:10]))
    best = max(worksheet_reviews, key=lambda review: (review.imported_rows, len([v for v in review.column_mapping.values() if v])), default=None)
    return WorkbookReview(
        file_name=name,
        supplier_name=detected_supplier,
        worksheet_names=excel.sheet_names,
        selected_worksheet=best.sheet_name if best else excel.sheet_names[0],
        header_row=best.header_row if best else 1,
        column_mapping=best.column_mapping if best else {},
        imported_rows=sum(w.imported_rows for w in worksheet_reviews),
        columns=best.columns if best else [],
        excluded_rows=sum(w.excluded_rows for w in worksheet_reviews),
        package_source=best.package_source if best else "Section headings",
        section_headings=[h for w in worksheet_reviews for h in w.section_headings][:10],
        worksheets=worksheet_reviews,
    )


def detect_supplier_name(file_name: str | Path) -> str:
    """Create a user-editable supplier name from an uploaded filename."""
    stem = Path(str(file_name)).stem
    cleaned = re.sub(r"[_\-]+", " ", stem)
    cleaned = FILENAME_NOISE_PATTERNS.sub(" ", cleaned)
    cleaned = re.sub(r"(?:^|\s)\d{3,}[a-z0-9]*(?=\s|$)", " ", cleaned, flags=re.I)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -_.,")
    return cleaned or "Supplier"


def read_excel_file(file: str | Path | BinaryIO, supplier_name: str | None = None, sheet_name: str | int | None = 0, header_row: int | None = None, column_mapping: dict[str, str | None] | None = None) -> pd.DataFrame:
    """Read one supplier spreadsheet and return commercial BOQ rows only."""
    supplier = supplier_name or detect_supplier_name(getattr(file, "name", str(file)))
    if sheet_name is None:
        return read_all_excel_sheets(file, supplier)
    if header_row is None or column_mapping is None:
        _rewind(file)
        excel = pd.ExcelFile(file, engine="openpyxl")
        header_row, _, detected = _detect_header(excel, sheet_name)
        column_mapping = column_mapping or detected
    rows, _, _ = _read_sheet(file, supplier, sheet_name, header_row or 0, column_mapping or {})
    return rows


def read_all_excel_sheets(file: str | Path | BinaryIO, supplier_name: str | None = None) -> pd.DataFrame:
    review = inspect_excel_file(file, supplier_name)
    frames = []
    for sheet in review.worksheets:
        frames.append(read_excel_file(file, review.supplier_name, sheet.sheet_name, sheet.header_row - 1, sheet.column_mapping))
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=COMMERCIAL_COLUMNS)


def _detect_header(excel: pd.ExcelFile, sheet: str | int) -> tuple[int, list[str], dict[str, str | None]]:
    preview = _read_preview_with_merged_headers(excel, sheet)
    best = (0, [], {}, -1, 1)
    for row_index in range(len(preview)):
        for header_rows, values in _logical_header_candidates(preview, row_index):
            mapping = detect_columns(values)
            numeric_cells = _numeric_cell_count(preview.iloc[row_index : row_index + header_rows])
            repeated_title_rows = _repeated_title_row_count(preview.iloc[row_index : row_index + header_rows])
            score = (
                len(mapping) * 3
                + int("description" in mapping)
                + int(any(k in mapping for k in ("quantity", "unit_rate", "total_amount")))
                + int("unit_rate" in mapping and "total_amount" in mapping)
                + (header_rows - 1 if "unit_rate" in mapping and "total_amount" in mapping else 0)
                - numeric_cells * 2
                - repeated_title_rows * 10
            )
            if score > best[3]:
                best = (row_index, [value for value in values if value], {c: mapping.get(c) for c in REQUIRED_COLUMNS}, score, header_rows)
    return best[0], best[1], best[2]


def _read_preview_with_merged_headers(excel: pd.ExcelFile, sheet: str | int) -> pd.DataFrame:
    """Read a preview while expanding merged cell values without saving the workbook."""
    source = excel._io  # pandas keeps the original path/file object here.
    _rewind(source)
    workbook = load_workbook(source, read_only=False, data_only=True)
    worksheet = workbook.worksheets[sheet] if isinstance(sheet, int) else workbook[sheet]
    max_row = min(worksheet.max_row, HEADER_SCAN_ROWS)
    max_column = worksheet.max_column
    values = [[worksheet.cell(row=row, column=column).value for column in range(1, max_column + 1)] for row in range(1, max_row + 1)]

    for merged_range in worksheet.merged_cells.ranges:
        top_value = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        if merged_range.min_row > max_row or top_value is None:
            continue
        for row in range(merged_range.min_row, min(merged_range.max_row, max_row) + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                values[row - 1][column - 1] = top_value

    workbook.close()
    _rewind(source)
    return pd.DataFrame(values).ffill(axis=1)


def _logical_header_candidates(preview: pd.DataFrame, row_index: int) -> Iterable[tuple[int, list[str]]]:
    for header_rows in range(MIN_HEADER_ROWS, MAX_HEADER_ROWS + 1):
        end = row_index + header_rows
        if end > len(preview):
            break
        header_block = preview.iloc[row_index:end]
        yield header_rows, _merge_stacked_headers(header_block)


def _merge_stacked_headers(header_block: pd.DataFrame) -> list[str]:
    logical_headers: list[str] = []
    for column in header_block.columns:
        parts: list[str] = []
        for value in header_block[column].tolist():
            cleaned = _clean_cell(value)
            if cleaned and cleaned not in parts:
                parts.append(cleaned)
        logical_headers.append(" ".join(parts))
    return _unique_headers(logical_headers)


def _numeric_cell_count(header_block: pd.DataFrame) -> int:
    count = 0
    for value in header_block.to_numpy().flatten():
        if _clean_cell(value) and pd.notna(pd.to_numeric(value, errors="coerce")):
            count += 1
    return count


def _repeated_title_row_count(header_block: pd.DataFrame) -> int:
    count = 0
    minimum_repeated_cells = max(2, len(header_block.columns) // 2)
    for _, row in header_block.iterrows():
        values = [_clean_cell(value) for value in row.tolist() if _clean_cell(value)]
        if len(values) >= minimum_repeated_cells and len(set(values)) == 1:
            count += 1
    return count


def _unique_headers(headers: Iterable[str]) -> list[str]:
    counts: dict[str, int] = {}
    unique: list[str] = []
    for index, header in enumerate(headers, start=1):
        name = header or f"Unnamed {index}"
        count = counts.get(name, 0)
        counts[name] = count + 1
        unique.append(f"{name}.{count}" if count else name)
    return unique


def _detect_header_rows_for_read(file: str | Path | BinaryIO, sheet: str | int, header_row: int, mapping: dict[str, str | None]) -> tuple[int, list[str]]:
    _rewind(file)
    excel = pd.ExcelFile(file, engine="openpyxl")
    preview = _read_preview_with_merged_headers(excel, sheet)
    expected_sources = {source for source in mapping.values() if source}
    best = (1, _merge_stacked_headers(preview.iloc[header_row : header_row + 1]), -1)
    for header_rows, columns in _logical_header_candidates(preview, header_row):
        matched = len(expected_sources.intersection(columns))
        mapping_score = len(detect_columns(columns))
        score = matched * 5 + mapping_score + (header_rows - 1 if matched else 0)
        if score > best[2]:
            best = (header_rows, columns, score)
    _rewind(file)
    return best[0], best[1]


def _read_sheet(file: str | Path | BinaryIO, supplier: str, sheet: str | int, header_row: int, mapping: dict[str, str | None]) -> tuple[pd.DataFrame, int, list[str]]:
    _rewind(file)
    header_rows, columns = _detect_header_rows_for_read(file, sheet, header_row, mapping)
    raw = pd.read_excel(file, engine="openpyxl", sheet_name=sheet, header=None, skiprows=header_row + header_rows)
    raw.columns = columns[: len(raw.columns)]
    raw = raw.ffill(axis=1)
    raw = raw.dropna(how="all")
    normalized = pd.DataFrame(index=raw.index)
    for canonical in REQUIRED_COLUMNS:
        source = _resolve_mapped_source(raw.columns, mapping.get(canonical))
        normalized[canonical] = raw[source] if source is not None else pd.NA
    package_col = _resolve_mapped_source(raw.columns, mapping.get("package")) or _find_package_column(raw.columns)
    normalized["package"] = raw[package_col] if package_col in raw.columns else pd.NA
    normalized["supplier"] = supplier
    normalized["file_name"] = Path(str(getattr(file, "name", str(file)))).name
    normalized["worksheet"] = str(sheet)
    normalized["item_no"] = normalized["item_no"].apply(lambda value: _clean_item_no(value) if pd.notna(value) else pd.NA)
    normalized["description"] = normalized["description"].apply(lambda value: _clean_cell(value) if pd.notna(value) else pd.NA)
    normalized["unit"] = normalized["unit"].apply(lambda value: _clean_cell(value) if pd.notna(value) else pd.NA)
    for col in ("quantity", "unit_rate", "total_amount"):
        normalized[col] = pd.to_numeric(normalized[col], errors="coerce")
    missing_total = normalized["total_amount"].isna() & normalized["quantity"].notna() & normalized["unit_rate"].notna()
    normalized.loc[missing_total, "total_amount"] = normalized.loc[missing_total, "quantity"] * normalized.loc[missing_total, "unit_rate"]
    imported, excluded, headings, current_package = [], 0, [], None
    for idx, row in normalized.iterrows():
        raw_text = " ".join(_clean_cell(v) for v in raw.loc[idx].tolist() if _clean_cell(v))
        heading = _section_heading(row, raw_text)
        if heading:
            current_package = heading
            headings.append(heading)
            excluded += 1
            continue
        row = row.copy()
        if pd.isna(row.get("package")) or not _clean_cell(row.get("package")):
            row["package"] = current_package
        if _is_valid_item(row, raw_text):
            imported.append(row)
        else:
            excluded += 1
    frame = pd.DataFrame(imported, columns=COMMERCIAL_COLUMNS)
    return frame.reset_index(drop=True), excluded, headings


def _is_valid_item(row: pd.Series, raw_text: str) -> bool:
    desc = _clean_cell(row.get("description"))
    if not desc or len(desc) < 3 or TOTAL_PATTERNS.search(desc) or NOTE_PATTERNS.search(desc):
        return False
    has_commercial = any(pd.notna(row.get(col)) for col in ("item_no", "quantity", "unit_rate", "total_amount"))
    return has_commercial and not TOTAL_PATTERNS.search(raw_text)


def _section_heading(row: pd.Series, raw_text: str) -> str | None:
    has_price_data = any(pd.notna(row.get(col)) for col in ("quantity", "unit_rate", "total_amount"))
    desc = _clean_cell(row.get("description"))
    item_no = _clean_cell(row.get("item_no"))
    if has_price_data or TOTAL_PATTERNS.search(raw_text):
        return None
    if desc and NON_PACKAGE_HEADINGS.match(desc):
        return None
    if desc and (SECTION_PATTERNS.search(desc) or (not item_no and len(desc.split()) <= 8 and not NOTE_PATTERNS.search(desc))):
        return desc
    return None


def _resolve_mapped_source(columns: Iterable[object], source: str | None) -> str | None:
    """Return the real worksheet column matching a stored mapping source.

    Only worksheet header names are considered here. The pandas row index is never
    inspected or promoted as a BOQ item number.
    """
    if not source:
        return None
    column_list = [str(column) for column in columns]
    if source in column_list:
        return source
    cleaned_source = re.sub(r"[^a-z0-9]+", " ", str(source).casefold()).strip()
    for column in column_list:
        cleaned_column = re.sub(r"[^a-z0-9]+", " ", str(column).casefold()).strip()
        if cleaned_column == cleaned_source:
            return column
    return None


def _find_package_column(columns: Iterable[object]) -> str | None:
    for col in columns:
        cleaned = re.sub(r"[^a-z0-9]+", " ", str(col).casefold()).strip()
        if cleaned in {"package", "section", "trade", "bill", "schedule", "package name", "section name"}:
            return str(col)
    return None


def _package_source(mapping: dict[str, str | None]) -> str:
    package = mapping.get("package")
    return str(package) if package else PACKAGE_FROM_HEADINGS


def _clean_cell(value: object) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _clean_item_no(value: object) -> str:
    """Preserve BOQ item numbers from Excel without pandas float artifacts."""
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _clean_cell(value)


def _rewind(file: str | Path | BinaryIO) -> None:
    if hasattr(file, "seek"):
        file.seek(0)


def read_reviewed_excels(files: Iterable[BinaryIO], reviews: list[WorkbookReview]) -> pd.DataFrame:
    frames = []
    for file, review in zip(files, reviews, strict=False):
        for sheet in (review.worksheets or [WorksheetReview(review.selected_worksheet, review.header_row, review.column_mapping, 0, 0, review.columns)]):
            if not any(sheet.column_mapping.get(col) for col in ("unit_rate", "total_amount")):
                continue
            frames.append(read_excel_file(file, review.supplier_name, sheet.sheet_name, sheet.header_row - 1, sheet.column_mapping))
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=COMMERCIAL_COLUMNS)


def read_uploaded_excels(files: Iterable[BinaryIO]) -> pd.DataFrame:
    reviews = [inspect_excel_file(file) for file in files]
    return read_reviewed_excels(files, reviews)


def read_sample_excels(directory: str | Path = "uploads") -> pd.DataFrame:
    paths = sample_workbooks(directory)
    frames = [read_excel_file(path) for path in paths]
    return pd.concat(frames, ignore_index=True)
