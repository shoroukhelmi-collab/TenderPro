from io import BytesIO

import pandas as pd
from openpyxl import Workbook

from modules.excel_reader import detect_supplier_name, inspect_excel_file, read_excel_file


def _book(rows_by_sheet, name="Acme Tender BOQ Rev 02.xlsx", merges=None):
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in rows_by_sheet.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
        for merge in (merges or {}).get(sheet_name, []):
            ws.merge_cells(merge)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    bio.name = name
    return bio


def test_title_rows_above_header_are_excluded_and_header_detected():
    file = _book({"BOQ": [
        ["Project Alpha"],
        ["Main Building Package"],
        ["Item No", "Description", "Unit", "Qty", "Rate", "Amount"],
        ["1.01", "Concrete works", "m3", 10, 20, 200],
    ]})

    review = inspect_excel_file(file)
    data = read_excel_file(file, review.supplier_name, "BOQ", review.worksheets[0].header_row - 1, review.worksheets[0].column_mapping)

    assert review.worksheets[0].header_row == 3
    assert len(data) == 1
    assert data.iloc[0]["description"] == "Concrete works"


def test_section_headings_are_inherited_as_package_not_items():
    file = _book({"BOQ": [
        ["Item", "Description", "Unit", "Quantity", "Unit Rate", "Total"],
        [None, "Section A - Civil Works", None, None, None, None],
        ["A1", "Excavation", "m3", 5, 3, 15],
    ]})

    data = read_excel_file(file)

    assert len(data) == 1
    assert data.iloc[0]["package"] == "Section A - Civil Works"


def test_subtotal_and_total_rows_are_excluded():
    file = _book({"BOQ": [
        ["Code", "Description", "Unit", "Qty", "Rate", "Amount"],
        ["1", "Doors", "ea", 2, 100, 200],
        [None, "Subtotal", None, None, None, 200],
        [None, "Grand Total", None, None, None, 200],
    ]})

    data = read_excel_file(file)

    assert data["description"].tolist() == ["Doors"]


def test_merged_cells_do_not_create_extra_items():
    file = _book({"BOQ": [
        ["Project title", None, None, None, None, None],
        ["Item No", "Description", "UOM", "Qty", "Rate", "Amount"],
        ["1", "Blockwork", "m2", 7, 4, 28],
    ]}, merges={"BOQ": ["A1:F1"]})

    data = read_excel_file(file)

    assert len(data) == 1
    assert data.iloc[0]["item_no"] == "1"


def test_multiple_worksheets_are_reviewed_and_read():
    file = _book({
        "Civil": [["Item", "Description", "Unit", "Qty", "Rate", "Amount"], ["C1", "Concrete", "m3", 1, 2, 2]],
        "MEP": [["Item", "Description", "Unit", "Qty", "Rate", "Amount"], ["M1", "Cable", "m", 3, 4, 12]],
    })

    review = inspect_excel_file(file)
    frames = [read_excel_file(file, review.supplier_name, ws.sheet_name, ws.header_row - 1, ws.column_mapping) for ws in review.worksheets]
    data = pd.concat(frames, ignore_index=True)

    assert {ws.sheet_name for ws in review.worksheets} == {"Civil", "MEP"}
    assert data["description"].tolist() == ["Concrete", "Cable"]


def test_supplier_name_comes_from_file_name_without_revision_noise():
    assert detect_supplier_name("/tmp/Acme Construction BOQ Rev 03 Final.xlsx") == "Acme Construction"


def test_invalid_descriptive_rows_are_excluded():
    file = _book({"BOQ": [
        ["Item No", "Description", "Unit", "Qty", "Rate", "Amount"],
        [None, "Note: rates include delivery", None, None, None, None],
        [None, "General description of works", None, None, None, None],
        ["1", "Valid item", "lot", 1, None, None],
    ]})

    data = read_excel_file(file)

    assert data["description"].tolist() == ["Valid item"]


def test_price_columns_detect_with_currency_suffixes():
    file = _book({"BOQ": [
        ["Ref", "Particulars", "UOM", "BOQ Qty", "Unit Rate (AED)", "Total Amount (AED)"],
        ["1", "Cable containment", "m", 10, 12.5, 125],
    ]})

    review = inspect_excel_file(file)
    mapping = review.worksheets[0].column_mapping
    data = read_excel_file(file, review.supplier_name, "BOQ", review.worksheets[0].header_row - 1, mapping)

    assert mapping["unit_rate"] == "Unit Rate (AED)"
    assert mapping["total_amount"] == "Total Amount (AED)"
    assert data.iloc[0]["unit_rate"] == 12.5
    assert data.iloc[0]["total_amount"] == 125


def test_multi_row_rate_and_amount_currency_headers_are_merged_and_read():
    file = _book({"Commercial BOQ": [
        ["Item No", "Description", "UOM", "Qty", "Rate", "Amount"],
        [None, None, None, None, "USD", "USD"],
        ["1", "Concrete foundations", "m3", 8, 100, 800],
    ]})

    review = inspect_excel_file(file)
    worksheet = review.worksheets[0]
    data = read_excel_file(file, review.supplier_name, "Commercial BOQ", worksheet.header_row - 1, worksheet.column_mapping)

    assert "Rate USD" in worksheet.columns
    assert "Amount USD" in worksheet.columns
    assert worksheet.column_mapping["unit_rate"] == "Rate USD"
    assert worksheet.column_mapping["total_amount"] == "Amount USD"
    assert data.iloc[0]["unit_rate"] == 100
    assert data.iloc[0]["total_amount"] == 800


def test_merged_multi_row_commercial_headers_are_expanded_without_modifying_workbook():
    file = _book({"BOQ": [
        [None, None, None, None, "Commercial", None],
        ["Item", "Particulars", "Unit", "Quantity", "Unit Price", "Line Total"],
        [None, None, None, None, "AED", "AED"],
        ["A1", "Blockwork", "m2", 12, 15, 180],
    ]}, merges={"BOQ": ["E1:F1"]})
    original_bytes = file.getvalue()

    review = inspect_excel_file(file)
    worksheet = review.worksheets[0]
    data = read_excel_file(file, review.supplier_name, "BOQ", worksheet.header_row - 1, worksheet.column_mapping)

    assert file.getvalue() == original_bytes
    assert worksheet.column_mapping["unit_rate"] == "Commercial Unit Price AED"
    assert worksheet.column_mapping["total_amount"] == "Commercial Line Total AED"
    assert data.iloc[0]["description"] == "Blockwork"
    assert data.iloc[0]["unit_rate"] == 15
    assert data.iloc[0]["total_amount"] == 180


def test_non_package_headings_are_not_inherited_as_packages():
    file = _book({"BOQ": [
        ["Item", "Description", "Unit", "Quantity", "Unit Rate", "Total"],
        [None, "Method of Measurements", None, None, None, None],
        [None, "General Notes", None, None, None, None],
        [None, "Specifications", None, None, None, None],
        [None, "Scope of Work", None, None, None, None],
        [None, "Section A - Civil Works", None, None, None, None],
        ["A1", "Excavation", "m3", 5, 3, 15],
    ]})

    data = read_excel_file(file)

    assert len(data) == 1
    assert data.iloc[0]["package"] == "Section A - Civil Works"


def test_supplier_name_removes_leading_project_number_revision_boq_and_extension():
    assert detect_supplier_name("3661 CMC04 FP Rev 2.xlsx") == "CMC04 FP"


def test_item_number_column_is_not_confused_with_item_description():
    file = _book({"BOQ": [
        ["Item Description", "UOM", "Qty", "Rate", "Amount"],
        ["Door set", "ea", 2, 100, 200],
    ]})

    review = inspect_excel_file(file)
    mapping = review.worksheets[0].column_mapping

    assert mapping.get("description") == "Item Description"
    assert mapping.get("item_no") is None


def test_code_column_populates_item_number_with_item_description_present():
    file = _book({"BOQ": [
        ["Code", "Item Description", "UOM", "Qty", "Rate", "Amount"],
        ["A-100", "Door set", "ea", 2, 100, 200],
    ]})

    data = read_excel_file(file)

    assert data.iloc[0]["item_no"] == "A-100"
    assert data.iloc[0]["description"] == "Door set"


def test_supplier_name_removes_generic_filename_noise_without_hardcoding_supplier():
    assert detect_supplier_name("987654 Alpha-Beta_Commercial BOQ Quotation R1.xlsx") == "Alpha Beta"


def test_reviewed_excels_use_explicit_supplier_group_name_for_multiple_files():
    from modules.excel_reader import read_reviewed_excels

    hvac = _book({"HVAC": [["Code", "Description", "Unit", "Qty", "Rate", "Amount"], ["H1", "Ductwork", "m", 2, 5, 10]]}, name="HVAC.xlsx")
    plumbing = _book({"Plumbing": [["Code", "Description", "Unit", "Qty", "Rate", "Amount"], ["P1", "Pipework", "m", 3, 7, 21]]}, name="Plumbing.xlsx")
    reviews = [inspect_excel_file(hvac, supplier_name="Hazek"), inspect_excel_file(plumbing, supplier_name="Hazek")]

    data = read_reviewed_excels([hvac, plumbing], reviews)

    assert data["supplier"].tolist() == ["Hazek", "Hazek"]
    assert data["file_name"].tolist() == ["HVAC.xlsx", "Plumbing.xlsx"]
    assert data["worksheet"].tolist() == ["HVAC", "Plumbing"]


def test_item_no_code_header_populates_item_number():
    file = _book({"BOQ": [
        ["ITEM / Item No / Code", "Description", "Unit", "Qty", "Rate", "Amount"],
        ["FA-01", "Smoke detector", "nr", 4, 10, 40],
    ]})

    data = read_excel_file(file)

    assert data.iloc[0]["item_no"] == "FA-01"


def test_uppercase_item_column_populates_item_no_without_dataframe_index():
    file = _book({"BOQ": [
        ["ITEM", "DESCRIPTION", "UNIT", "QTY", "RATE", "AMOUNT"],
        [5, "50mm", "m", 50, 71, 3550],
    ]})

    data = read_excel_file(file)

    assert data.index.tolist() == [0]
    assert data.iloc[0]["item_no"] == "5"
    assert data.iloc[0]["description"] == "50mm"
    assert data.iloc[0]["unit"] == "m"
    assert data.iloc[0]["quantity"] == 50
    assert data.iloc[0]["unit_rate"] == 71
    assert data.iloc[0]["total_amount"] == 3550
